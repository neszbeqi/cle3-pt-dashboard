"""
CLE3 PT Daily Report Generator
Runs automatically at 7 PM via Task Scheduler.

Fetches all 3 shifts for the current day:
  - Night Shift (6p-6a)  → yesterday's date (shift ended 6 AM today)
  - Day Shift  (6a-6p)   → today's date     (shift ended 6 PM today)
  - All Day               → today's date     (6 AM – 6 PM)

For each shift:
  • Saves a formatted Excel daily report to Desktop (Night Shift only — primary report)
  • Appends to CLE3_PT_History.xlsx
  • Updates associate history JSON

After all 3 fetches, rebuilds CLE3_PT_Trends.xlsx automatically.
"""
import os, sys, traceback, json as _json
from datetime import datetime, timedelta

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
import fclm, processor, history

from openpyxl import Workbook
from openpyxl.styles import (PatternFill, Font, Alignment, Border, Side, GradientFill)
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.chart.series import DataPoint
from openpyxl.utils import get_column_letter

# ── Color palette (matches dashboard) ─────────────────────────────────────────
BG_DARK   = "0F1117"; CARD = "1A1D27"; ACCENT = "FF9900"; WHITE = "FFFFFF"
MUTED     = "7A7F96"; C_RED = "E05260"; C_ORA = "F59E0B"
C_LGR     = "4ADE80"; C_DGR = "16A34A"; C_HEADER = "12141E"; C_ROW_ALT = "1E2130"

def pt_color(pt):
    if pt is None: return MUTED
    if pt < 80:    return C_RED
    if pt < 85:    return C_ORA
    if pt < 90:    return C_LGR
    return C_DGR

def pt_str(pt):
    return f"{pt:.1f}%" if pt is not None else "—"

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def font(bold=False, color=WHITE, size=11):
    return Font(bold=bold, color=color, name="Segoe UI", size=size)

def center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def left():
    return Alignment(horizontal="left", vertical="center")

def thin_border():
    s = Side(style="thin", color="2A2D3A")
    return Border(left=s, right=s, top=s, bottom=s)

def set_cell(ws, row, col, value, bold=False, color=WHITE, bg=None,
             align="center", size=11, number_format=None):
    c = ws.cell(row=row, column=col, value=value)
    c.font      = Font(bold=bold, color=color, name="Segoe UI", size=size)
    c.alignment = center() if align == "center" else left()
    if bg: c.fill = fill(bg)
    c.border = thin_border()
    if number_format: c.number_format = number_format
    return c

# ── Sheet builders ─────────────────────────────────────────────────────────────

def build_summary(wb, data, date_str, shift_label):
    ws = wb.active
    ws.title = "Summary"
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 22
    ws.row_dimensions[1].height = 40

    ws.merge_cells("A1:B1")
    c = ws["A1"]
    c.value     = f"CLE3 Productive Time  ·  {shift_label}  ·  {date_str}"
    c.font      = Font(bold=True, color=ACCENT, name="Segoe UI", size=16)
    c.fill      = fill(CARD)
    c.alignment = center()

    ws.merge_cells("A2:B2")
    c = ws["A2"]
    c.value     = f"Generated {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    c.font      = Font(color=MUTED, name="Segoe UI", size=10)
    c.fill      = fill(CARD)
    c.alignment = center()

    rows = [
        ("Overall PT%",        pt_str(data["overall_pt"]),    pt_color(data["overall_pt"])),
        ("Total Associates",   data["aa_count"],               WHITE),
        ("Flagged  (<84%)",    data["flagged"],                C_ORA if data["flagged"] else C_DGR),
        ("Above 90%",          data["above_90"],               C_DGR),
        ("Managers",           len(data["managers"]),          WHITE),
        ("Total Inferred Hrs", f"{data['total_inferred']:.1f}", C_ORA),
        ("Total Hours",        f"{data['total_hours']:.1f}",   WHITE),
    ]
    for i, (label, value, color) in enumerate(rows):
        r = i + 4
        ws.row_dimensions[r].height = 28
        set_cell(ws, r, 1, label, bold=False, color=MUTED, bg=C_HEADER, align="left", size=11)
        set_cell(ws, r, 2, value, bold=True,  color=color, bg=CARD,     align="center", size=13)

    ws.row_dimensions[12].height = 28
    ws.merge_cells("A11:B11")
    c = ws["A11"]
    c.value = "PT% Distribution"
    c.font  = Font(bold=True, color=ACCENT, name="Segoe UI", size=12)
    c.fill  = fill(CARD); c.alignment = center()

    valid = [a for a in data["associates"] if a["pt"] is not None]
    tiers = [
        ("≥ 90%",    sum(1 for a in valid if a["pt"] >= 90),       C_DGR),
        ("85 – 89%", sum(1 for a in valid if 85 <= a["pt"] < 90),  C_LGR),
        ("80 – 84%", sum(1 for a in valid if 80 <= a["pt"] < 85),  C_ORA),
        ("< 80%",    sum(1 for a in valid if a["pt"] < 80),         C_RED),
    ]
    for i, (label, count, color) in enumerate(tiers):
        r = i + 12
        ws.row_dimensions[r].height = 26
        set_cell(ws, r, 1, label, bold=False, color=MUTED, bg=C_HEADER, align="left")
        set_cell(ws, r, 2, count, bold=True,  color=color, bg=CARD,     align="center", size=13)

    pie = PieChart()
    pie.title = "PT% Tier Distribution"; pie.style = 10
    pie.width = 14; pie.height = 10

    for i, (label, count, _) in enumerate(tiers):
        ws.cell(row=i+12, column=4, value=label)
        ws.cell(row=i+12, column=5, value=count)

    labels   = Reference(ws, min_col=4, min_row=12, max_row=15)
    data_ref = Reference(ws, min_col=5, min_row=12, max_row=15)
    pie.add_data(data_ref); pie.set_categories(labels)

    slice_colors = [C_DGR, C_LGR, C_ORA, C_RED]
    for idx, hex_c in enumerate(slice_colors):
        pt_data = DataPoint(idx=idx)
        pt_data.graphicalProperties.solidFill = hex_c
        pie.series[0].dPt.append(pt_data)

    ws.add_chart(pie, "D4")
    ws["D4"].fill = fill(CARD)


def build_am_rankings(wb, data):
    ws = wb.create_sheet("AM Rankings")
    ws.sheet_view.showGridLines = False
    cols   = ["Rank","Area Manager","PT%","Associates","Flagged","Inferred Hrs","Total Hrs"]
    widths = [8, 30, 12, 12, 10, 14, 12]
    for i, (col, w) in enumerate(zip(cols, widths)):
        ws.column_dimensions[get_column_letter(i+1)].width = w

    ws.row_dimensions[1].height = 36
    ws.merge_cells("A1:G1")
    c = ws["A1"]
    c.value = "AM Rankings — Productive Time"
    c.font  = Font(bold=True, color=ACCENT, name="Segoe UI", size=14)
    c.fill  = fill(CARD); c.alignment = center()

    ws.row_dimensions[2].height = 26
    for i, col in enumerate(cols):
        set_cell(ws, 2, i+1, col, bold=True, color=WHITE, bg=C_HEADER, size=10)

    for rank, mg in enumerate(data["managers"]):
        r  = rank + 3
        bg = CARD if rank % 2 == 0 else C_ROW_ALT
        color = pt_color(mg["pt"])
        ws.row_dimensions[r].height = 22
        medal = ["🥇","🥈","🥉"][rank] if rank < 3 else str(rank+1)
        set_cell(ws, r, 1, medal,               bg=bg, align="center")
        set_cell(ws, r, 2, mg["name"],          bg=bg, align="left")
        set_cell(ws, r, 3, pt_str(mg["pt"]),    bg=bg, color=color, bold=True)
        set_cell(ws, r, 4, mg["aa_count"],      bg=bg)
        set_cell(ws, r, 5, mg["flagged"],       bg=bg, color=C_ORA if mg["flagged"] else C_DGR)
        set_cell(ws, r, 6, round(mg["inferred"],2), bg=bg, number_format="0.00")
        set_cell(ws, r, 7, round(mg["total"],2),    bg=bg, number_format="0.00")

    chart_start_row = 3
    for rank, mg in enumerate(data["managers"]):
        r = rank + chart_start_row
        ws.cell(row=r, column=9,  value=mg["name"])
        ws.cell(row=r, column=10, value=round(mg["pt"] or 0, 1))

    n   = len(data["managers"])
    bar = BarChart(); bar.type = "bar"; bar.grouping = "clustered"
    bar.title = "Manager PT% Rankings"; bar.style = 10
    bar.width = 22; bar.height = max(10, n * 0.45)
    data_ref = Reference(ws, min_col=10, min_row=chart_start_row, max_row=chart_start_row+n-1)
    cats     = Reference(ws, min_col=9,  min_row=chart_start_row, max_row=chart_start_row+n-1)
    bar.add_data(data_ref); bar.set_categories(cats)
    bar.series[0].graphicalProperties.solidFill      = ACCENT
    bar.series[0].graphicalProperties.line.solidFill = ACCENT
    bar.y_axis.scaling.min = 0; bar.y_axis.scaling.max = 100
    ws.add_chart(bar, "A" + str(n + 5))


def build_associates(wb, data):
    ws = wb.create_sheet("All Associates")
    ws.sheet_view.showGridLines = False
    cols   = ["Rank","Badge ID","Name","Manager","PT%","Inferred Hrs","Total Hrs"]
    widths = [8, 14, 28, 28, 12, 14, 12]
    for i, (col, w) in enumerate(zip(cols, widths)):
        ws.column_dimensions[get_column_letter(i+1)].width = w

    ws.row_dimensions[1].height = 36
    ws.merge_cells("A1:G1")
    c = ws["A1"]
    c.value = "All Associates — Sorted by PT% (lowest first)"
    c.font  = Font(bold=True, color=ACCENT, name="Segoe UI", size=14)
    c.fill  = fill(CARD); c.alignment = center()

    ws.row_dimensions[2].height = 26
    for i, col in enumerate(cols):
        set_cell(ws, 2, i+1, col, bold=True, color=WHITE, bg=C_HEADER, size=10)

    sorted_aas = sorted(data["associates"], key=lambda a: a["pt"] or 99)
    for rank, aa in enumerate(sorted_aas):
        r  = rank + 3; bg = CARD if rank % 2 == 0 else C_ROW_ALT
        ws.row_dimensions[r].height = 20
        color = pt_color(aa["pt"])
        set_cell(ws, r, 1, rank+1,            bg=bg)
        set_cell(ws, r, 2, aa["id"],          bg=bg, align="left")
        set_cell(ws, r, 3, aa["name"],        bg=bg, align="left")
        set_cell(ws, r, 4, aa["manager"],     bg=bg, align="left")
        set_cell(ws, r, 5, pt_str(aa["pt"]),  bg=bg, color=color, bold=True)
        set_cell(ws, r, 6, aa["inferred"],    bg=bg, number_format="0.00")
        set_cell(ws, r, 7, aa["total"],       bg=bg, number_format="0.00")


def build_flagged(wb, data, threshold=84):
    ws = wb.create_sheet(f"Flagged  (<{threshold}%)")
    ws.sheet_view.showGridLines = False
    cols   = ["Rank","Badge ID","Name","Manager","PT%","Gap to 84%","Inferred Hrs","Total Hrs"]
    widths = [8, 14, 28, 28, 12, 12, 14, 12]
    for i, (col, w) in enumerate(zip(cols, widths)):
        ws.column_dimensions[get_column_letter(i+1)].width = w

    ws.row_dimensions[1].height = 36
    ws.merge_cells("A1:H1")
    c = ws["A1"]
    c.value = f"Flagged Associates  ·  PT% below {threshold}%"
    c.font  = Font(bold=True, color=C_RED, name="Segoe UI", size=14)
    c.fill  = fill(CARD); c.alignment = center()

    ws.row_dimensions[2].height = 26
    for i, col in enumerate(cols):
        set_cell(ws, 2, i+1, col, bold=True, color=WHITE, bg=C_HEADER, size=10)

    flagged = sorted(
        [a for a in data["associates"] if a["pt"] is not None and a["pt"] < threshold],
        key=lambda a: a["pt"])

    if not flagged:
        ws.merge_cells("A3:H3")
        c = ws["A3"]
        c.value = f"✓  No associates below {threshold}% — great shift!"
        c.font  = Font(bold=True, color=C_DGR, name="Segoe UI", size=12)
        c.fill  = fill(CARD); c.alignment = center()
        return

    for rank, aa in enumerate(flagged):
        r     = rank + 3; bg = CARD if rank % 2 == 0 else C_ROW_ALT
        color = pt_color(aa["pt"]); gap = round(threshold - aa["pt"], 1)
        ws.row_dimensions[r].height = 20
        set_cell(ws, r, 1, rank+1,           bg=bg)
        set_cell(ws, r, 2, aa["id"],         bg=bg, align="left")
        set_cell(ws, r, 3, aa["name"],       bg=bg, align="left")
        set_cell(ws, r, 4, aa["manager"],    bg=bg, align="left")
        set_cell(ws, r, 5, pt_str(aa["pt"]),bg=bg, color=color, bold=True)
        set_cell(ws, r, 6, f"−{gap}%",      bg=bg, color=C_ORA)
        set_cell(ws, r, 7, aa["inferred"],   bg=bg, number_format="0.00")
        set_cell(ws, r, 8, aa["total"],      bg=bg, number_format="0.00")

    chart_col = 10
    for rank, aa in enumerate(flagged):
        ws.cell(row=rank+3, column=chart_col,   value=aa["name"])
        ws.cell(row=rank+3, column=chart_col+1, value=round(aa["pt"] or 0, 1))

    n   = len(flagged)
    bar = BarChart(); bar.type = "bar"; bar.grouping = "clustered"
    bar.title = f"Flagged Associates PT%  (below {threshold}%)"
    bar.style = 10; bar.width = 22; bar.height = max(10, n * 0.45)
    bar.y_axis.scaling.min = 0; bar.y_axis.scaling.max = 100
    data_ref = Reference(ws, min_col=chart_col+1, min_row=3, max_row=3+n-1)
    cats     = Reference(ws, min_col=chart_col,   min_row=3, max_row=3+n-1)
    bar.add_data(data_ref); bar.set_categories(cats)
    bar.series[0].graphicalProperties.solidFill      = C_ORA
    bar.series[0].graphicalProperties.line.solidFill = C_ORA
    ws.add_chart(bar, "A" + str(n + 5))


# ── Associate history helper ───────────────────────────────────────────────────

def _update_aa_history(date_str, shift_name, data, log):
    aa_hist_path = os.path.join(os.path.expanduser("~"), ".pt_dashboard", "associate_history.json")
    try:
        if os.path.exists(aa_hist_path):
            with open(aa_hist_path) as _hf:
                aa_hist = _json.load(_hf)
        else:
            aa_hist = {}
        for aa in data["associates"]:
            b = aa["id"]
            if b not in aa_hist:
                aa_hist[b] = {"name": aa["name"], "manager": aa["manager"], "entries": []}
            aa_hist[b]["entries"] = [e for e in aa_hist[b]["entries"]
                                     if not (e["date"] == date_str and e["shift"] == shift_name)]
            aa_hist[b]["entries"].append({"date": date_str, "shift": shift_name, "pt": aa["pt"]})
            aa_hist[b]["entries"].sort(key=lambda e: e["date"], reverse=True)
            aa_hist[b]["entries"] = aa_hist[b]["entries"][:60]
        with open(aa_hist_path, "w") as _hf:
            _json.dump(aa_hist, _hf)
        log(f"Associate history updated ({shift_name}): {aa_hist_path}")
    except Exception as _e:
        log(f"Warning: could not update associate history ({shift_name}): {_e}")


# ── Fetch one shift + update history ──────────────────────────────────────────

def fetch_shift(date_str, shift_name, shift_label, warehouse, log):
    """Fetch one shift from FCLM, append to history, update AA history."""
    log(f"--- Fetching {shift_label} for {date_str} ---")
    result = fclm.fetch(date_str, shift_name, warehouse, status_cb=log)
    if not result["ok"]:
        log(f"FCLM fetch failed ({shift_label}): {result['error']}")
        return None

    data = processor.process(result["rows"])
    log(f"  {shift_label}: {data['aa_count']} associates, PT={data['overall_pt']}%")

    hist_path = history.update(date_str, shift_label, data)
    log(f"  History updated: {hist_path}")

    _update_aa_history(date_str, shift_name, data, log)
    return data


# ── Main ───────────────────────────────────────────────────────────────────────

def run():
    today     = datetime.today()
    yesterday = today - timedelta(days=1)
    today_str = today.strftime("%Y-%m-%d")
    yest_str  = yesterday.strftime("%Y-%m-%d")
    warehouse = "CLE3"

    log_path = os.path.join(os.path.expanduser("~"), ".pt_dashboard", "report_log.txt")

    def log(msg):
        print(msg)
        with open(log_path, "a") as f:
            f.write(f"[{datetime.now().strftime('%H:%M:%S')}] {msg}\n")

    log(f"\n=== Report run started: {today_str} ===")

    # ── 1. Night Shift (yesterday 6p → today 6a) ──────────────────────────────
    night_data = fetch_shift(
        yest_str,
        "Night Shift (6p-6a)",
        "Night Shift (6p – 6a)",
        warehouse, log)

    # ── 2. Day Shift (today 6a → 6p) ─────────────────────────────────────────
    fetch_shift(
        today_str,
        "Day Shift  (6a-6p)",
        "Day Shift (6a – 6p)",
        warehouse, log)

    # ── 3. All Day (today 6a → 6p aggregated) ────────────────────────────────
    fetch_shift(
        today_str,
        "All Day",
        "All Day",
        warehouse, log)

    # ── Save nightly dashboard snapshot (Night Shift only, used by dashboard) ─
    if night_data:
        snapshot_path = os.path.join(os.path.expanduser("~"), ".pt_dashboard", "last_report.json")
        with open(snapshot_path, "w") as _sf:
            _json.dump({
                "date_str":     yest_str,
                "shift_name":   "Night Shift (6p-6a)",
                "shift_label":  "Night Shift (6p – 6a)",
                "warehouse":    warehouse,
                "generated_at": datetime.now().strftime("%Y-%m-%d %H:%M"),
                "data":         night_data,
            }, _sf)
        log(f"Dashboard snapshot saved: {snapshot_path}")

        # ── Build formatted daily Excel report (Night Shift) ──────────────────
        wb = Workbook()
        build_summary(wb, night_data, yest_str, "Night Shift (6p – 6a)")
        build_am_rankings(wb, night_data)
        build_associates(wb, night_data)
        build_flagged(wb, night_data)

        desktop  = os.path.join(os.path.expanduser("~"), "Desktop")
        filename = f"CLE3_PT_{yest_str}.xlsx"
        out_path = os.path.join(desktop, filename)
        try:
            wb.save(out_path)
            log(f"Daily report saved: {out_path}")
            os.startfile(out_path)
        except PermissionError:
            alt_path = os.path.join(desktop, f"CLE3_PT_{yest_str}_new.xlsx")
            wb.save(alt_path)
            log(f"Daily report saved (original locked, used alt): {alt_path}")
            os.startfile(alt_path)

    # ── Rebuild 3-tab trends workbook ─────────────────────────────────────────
    log("Rebuilding CLE3_PT_Trends.xlsx ...")
    try:
        import importlib.util, pathlib
        # load build_trends from its installed location alongside this script
        _candidates = [
            os.path.join(os.path.dirname(os.path.abspath(__file__)), "build_trends.py"),
            r"C:\Users\neszbeqi\.aki\aki_workspace\build_trends.py",
        ]
        _bt_path = next((p for p in _candidates if os.path.exists(p)), None)
        if _bt_path:
            spec = importlib.util.spec_from_file_location("build_trends", _bt_path)
            bt   = importlib.util.module_from_spec(spec)
            spec.loader.exec_module(bt)
            bt.main()
            log(f"Trends rebuilt: {bt.OUT_PATH}")
        else:
            log("Warning: build_trends.py not found — trends not rebuilt")
    except Exception as _te:
        log(f"Warning: trends rebuild failed: {_te}")

    log("=== Run complete ===")


if __name__ == "__main__":
    try:
        run()
    except Exception:
        log_path = os.path.join(os.path.expanduser("~"), ".pt_dashboard", "report_log.txt")
        with open(log_path, "a") as f:
            f.write(traceback.format_exc())
        raise
