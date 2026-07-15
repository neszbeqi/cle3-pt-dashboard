"""
CLE3 PT History — master workbook that grows one row per night shift.
Called automatically by generate_report.py after every 7 PM run.

Workbook: CLE3_PT_History.xlsx  (Desktop)
  Sheet "Daily Summary"  — one row per shift + charts
  Sheet "Manager Trends" — every manager's stats per night (long format)
  Sheet "Charts"         — large standalone charts for easy reading
"""
import os
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.series import DataPoint
from openpyxl.chart.label import DataLabelList
from openpyxl.utils import get_column_letter

_REPORTS_DIR = os.path.join(os.path.expanduser("~"), "Documents", "CLE3_Reports")
os.makedirs(_REPORTS_DIR, exist_ok=True)
HISTORY_PATH = os.path.join(_REPORTS_DIR, "CLE3_PT_History.xlsx")

# ── Palette ────────────────────────────────────────────────────────────────────
CARD      = "1A1D27"; C_HEADER = "12141E"; C_ROW_ALT = "1E2130"
ACCENT    = "FF9900"; WHITE    = "FFFFFF"; MUTED     = "7A7F96"
C_RED     = "E05260"; C_ORA    = "F59E0B"; C_LGR     = "4ADE80"; C_DGR = "16A34A"
TARGET_COLOR = "FF4444"   # dashed reference line at 84%

def _fill(h): return PatternFill("solid", fgColor=h)
def _border():
    s = Side(style="thin", color="2A2D3A")
    return Border(left=s, right=s, top=s, bottom=s)

def _pt_color(pt):
    if pt is None: return MUTED
    if pt < 80:    return C_RED
    if pt < 85:    return C_ORA
    if pt < 90:    return C_LGR
    return C_DGR

def _sc(ws, r, c, v, bold=False, col=WHITE, bg=None, align="center", fmt=None, size=11):
    cell = ws.cell(row=r, column=c, value=v)
    cell.font      = Font(bold=bold, color=col, name="Segoe UI", size=size)
    cell.alignment = Alignment(horizontal=align, vertical="center")
    if bg: cell.fill = _fill(bg)
    cell.border = _border()
    if fmt: cell.number_format = fmt
    return cell

# ── Column headers ─────────────────────────────────────────────────────────────
SUMMARY_COLS  = ["Date","Shift","Overall PT%","Associates","Flagged (<84%)","Above 90%","Inferred Hrs","Total Hrs","Target 84%"]
SUMMARY_W     = [14,    22,      13,           12,          15,              11,          13,            11,          0]   # last col hidden (chart helper)
MANAGER_COLS  = ["Date","Shift","Manager","PT%","Associates","Flagged","Inferred Hrs","Total Hrs"]
MANAGER_W     = [14,    22,     30,        12,   12,          10,       13,            11]


def _init_summary(ws):
    ws.title = "Daily Summary"
    ws.sheet_view.showGridLines = False
    for i, (h, w) in enumerate(zip(SUMMARY_COLS, SUMMARY_W)):
        col_ltr = get_column_letter(i + 1)
        ws.column_dimensions[col_ltr].width  = w if w > 0 else 0.1
        ws.column_dimensions[col_ltr].hidden = (w == 0)
    ws.row_dimensions[1].height = 28
    for i, h in enumerate(SUMMARY_COLS):
        if i < len(SUMMARY_W) - 1:   # skip hidden helper col
            _sc(ws, 1, i+1, h, bold=True, bg=C_HEADER, size=10)


def _init_manager(ws):
    ws.title = "Manager Trends"
    ws.sheet_view.showGridLines = False
    for i, (h, w) in enumerate(zip(MANAGER_COLS, MANAGER_W)):
        ws.column_dimensions[get_column_letter(i+1)].width = w
    ws.row_dimensions[1].height = 28
    for i, h in enumerate(MANAGER_COLS):
        _sc(ws, 1, i+1, h, bold=True, bg=C_HEADER, size=10)


# ── Chart builder ──────────────────────────────────────────────────────────────

def _make_pt_chart(ws_sum, n_data, width=28, height=16):
    """
    Column chart: one bar per night, colored by PT% tier.
    Overlaid with a dashed red line at the 84% target.
    Data labels on every bar showing exact value.
    Y-axis starts at 60 so differences are easy to see.
    """
    # --- Bar series: Overall PT% ---
    bar = BarChart()
    bar.type      = "col"
    bar.grouping  = "clustered"
    bar.title     = "CLE3 Night Shift  ·  Overall PT% per Date"
    bar.style     = 2          # clean minimal Excel style
    bar.width     = width
    bar.height    = height
    bar.y_axis.title      = "Productive Time %"
    bar.x_axis.title      = "Shift Date"
    bar.y_axis.scaling.min = 60
    bar.y_axis.scaling.max = 100
    bar.y_axis.majorUnit   = 5
    bar.y_axis.numFmt      = '0"%"'
    bar.x_axis.tickLblSkip = 1
    bar.plot_area.spPr     = None

    pt_ref  = Reference(ws_sum, min_col=3, min_row=1, max_row=1 + n_data)
    bar.add_data(pt_ref, titles_from_data=True)
    bar.set_categories(Reference(ws_sum, min_col=1, min_row=2, max_row=1 + n_data))

    # Color each bar by PT% tier
    for i in range(n_data):
        pt_val = ws_sum.cell(row=i + 2, column=3).value
        color  = _pt_color(pt_val)
        dp = DataPoint(idx=i)
        dp.graphicalProperties.solidFill      = color
        dp.graphicalProperties.line.solidFill = color
        bar.series[0].dPt.append(dp)

    # Data labels on top of each bar
    dLbls = DataLabelList()
    dLbls.showVal      = True
    dLbls.showLegendKey= False
    dLbls.showCatName  = False
    dLbls.showSerName  = False
    dLbls.position     = "outEnd"   # above the bar
    bar.series[0].dLbls = dLbls
    bar.series[0].graphicalProperties.line.solidFill = "FFFFFF"  # no bar outline

    # --- Line series: 84% target (column 9, hidden) ---
    line = LineChart()
    tgt_ref = Reference(ws_sum, min_col=9, min_row=1, max_row=1 + n_data)
    line.add_data(tgt_ref, titles_from_data=True)
    line.set_categories(Reference(ws_sum, min_col=1, min_row=2, max_row=1 + n_data))

    s = line.series[0]
    s.graphicalProperties.line.solidFill = TARGET_COLOR
    s.graphicalProperties.line.width     = 22000   # ~1.75pt
    s.graphicalProperties.line.dashDot   = "dash"
    s.marker.symbol = "none"
    # No labels on target line
    tgt_lbls = DataLabelList()
    tgt_lbls.showVal = False
    s.dLbls = tgt_lbls

    # Combine bar + target line into one chart object
    bar += line
    return bar


def _make_flagged_chart(ws_sum, n_data, width=28, height=13):
    """Column chart: flagged count per night, bars colored amber/red."""
    bar = BarChart()
    bar.type     = "col"
    bar.grouping = "clustered"
    bar.title    = "Associates Flagged Below 84%  ·  Per Night"
    bar.style    = 2
    bar.width    = width
    bar.height   = height
    bar.y_axis.title = "# Associates"
    bar.x_axis.title = "Shift Date"
    bar.y_axis.scaling.min = 0
    bar.x_axis.tickLblSkip = 1

    flag_ref = Reference(ws_sum, min_col=5, min_row=1, max_row=1 + n_data)
    bar.add_data(flag_ref, titles_from_data=True)
    bar.set_categories(Reference(ws_sum, min_col=1, min_row=2, max_row=1 + n_data))

    for i in range(n_data):
        val = ws_sum.cell(row=i + 2, column=5).value or 0
        color = C_RED if val > 40 else C_ORA
        dp = DataPoint(idx=i)
        dp.graphicalProperties.solidFill      = color
        dp.graphicalProperties.line.solidFill = color
        bar.series[0].dPt.append(dp)

    dLbls = DataLabelList()
    dLbls.showVal       = True
    dLbls.showLegendKey = False
    dLbls.showCatName   = False
    dLbls.showSerName   = False
    dLbls.position      = "outEnd"
    bar.series[0].dLbls  = dLbls
    bar.series[0].graphicalProperties.line.solidFill = "FFFFFF"
    return bar


def _rebuild_charts(wb):
    ws_sum   = wb["Daily Summary"]
    ws_chart = wb["Charts"]
    ws_sum._charts.clear()
    ws_chart._charts.clear()

    n = ws_sum.max_row - 1   # rows of actual data (excludes header)
    if n < 1:
        return

    # Ensure target column (col 9) is filled for every data row
    ws_sum.cell(1, 9).value = "Target 84%"
    for r in range(2, n + 2):
        ws_sum.cell(r, 9).value = 84

    pt_chart   = _make_pt_chart(ws_sum, n)
    flag_chart = _make_flagged_chart(ws_sum, n)

    # Place on Charts sheet — large and easy to read
    ws_chart.add_chart(pt_chart,   "A1")
    ws_chart.add_chart(flag_chart, "A35")

    # Smaller copies embedded below data on Daily Summary
    pt_small   = _make_pt_chart(ws_sum, n, width=22, height=14)
    flag_small = _make_flagged_chart(ws_sum, n, width=22, height=11)
    anchor = n + 4
    ws_sum.add_chart(pt_small,   f"A{anchor}")
    ws_sum.add_chart(flag_small, f"A{anchor + 22}")


# ── Public API ─────────────────────────────────────────────────────────────────

def update(date_str, shift_label, data):
    """Append a shift result to the master history workbook and refresh charts."""
    if os.path.exists(HISTORY_PATH):
        import io
        with open(HISTORY_PATH, 'rb') as _fh:
            _data = _fh.read()
        wb = load_workbook(io.BytesIO(_data))
        ws_sum = wb["Daily Summary"]
        ws_mgr = wb["Manager Trends"]
    else:
        wb = Workbook()
        ws_sum = wb.active
        _init_summary(ws_sum)
        ws_mgr = wb.create_sheet()
        _init_manager(ws_mgr)
        ws_chart = wb.create_sheet()
        ws_chart.title = "Charts"
        ws_chart.sheet_view.showGridLines = False

    # Remove duplicate rows for this date+shift if re-running
    key = f"{date_str}|{shift_label}"
    for ws in (ws_sum, ws_mgr):
        to_del = [r for r in range(2, ws.max_row + 1)
                  if f"{ws.cell(r,1).value}|{ws.cell(r,2).value}" == key]
        for r in reversed(to_del):
            ws.delete_rows(r)

    # ── Daily Summary row ─────────────────────────────────────────────────────
    r  = ws_sum.max_row + 1
    bg = CARD if r % 2 == 0 else C_ROW_ALT
    ws_sum.row_dimensions[r].height = 22
    summary_vals = [
        date_str, shift_label,
        data["overall_pt"], data["aa_count"],
        data["flagged"],    data["above_90"],
        round(data["total_inferred"], 1),
        round(data["total_hours"],    1),
        84,   # hidden target column for chart reference line
    ]
    for c, val in enumerate(summary_vals, 1):
        color = _pt_color(data["overall_pt"]) if c == 3 else (
                C_ORA if c == 5 and data["flagged"] > 0 else
                C_DGR if c == 5 else WHITE)
        _sc(ws_sum, r, c, val, col=color, bg=bg,
            align="left" if c <= 2 else "center")

    # ── Manager Trends rows ───────────────────────────────────────────────────
    for mg in data["managers"]:
        r2 = ws_mgr.max_row + 1
        bg2 = CARD if r2 % 2 == 0 else C_ROW_ALT
        ws_mgr.row_dimensions[r2].height = 20
        for c, val in enumerate([
            date_str, shift_label, mg["name"], mg["pt"],
            mg["aa_count"], mg["flagged"],
            round(mg["inferred"], 2), round(mg["total"], 2)
        ], 1):
            color = _pt_color(mg["pt"]) if c == 4 else WHITE
            _sc(ws_mgr, r2, c, val, col=color, bg=bg2,
                align="left" if c <= 3 else "center")

    _rebuild_charts(wb)
    import tempfile, shutil, time
    tmp = HISTORY_PATH + '.tmp'
    wb.save(tmp)
    for _attempt in range(6):
        try:
            if os.path.exists(HISTORY_PATH):
                os.remove(HISTORY_PATH)
            shutil.move(tmp, HISTORY_PATH)
            break
        except PermissionError:
            if _attempt == 5: raise
            time.sleep(1)
    return HISTORY_PATH
