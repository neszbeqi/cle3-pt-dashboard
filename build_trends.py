"""
CLE3 PT Trends — 3-tab Excel workbook.
  Tab 1 — Overall PT% Trend    : 3 lines (Night / Day / All Day) + 84% target
  Tab 2 — AM PT% Trends        : one chart per shift, stacked vertically
  Tab 3 — Flagged Count Trend  : 3 lines (Night / Day / All Day)

Auto-rebuilt every night at 7 PM by generate_report.py.
"""
import openpyxl
from openpyxl.chart import LineChart, Reference
from openpyxl.chart.series import SeriesLabel
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from collections import defaultdict
import datetime

SRC_PATH = r'C:\Users\neszbeqi\Documents\CLE3_Reports\CLE3_PT_History.xlsx'
OUT_PATH = r'C:\Users\neszbeqi\Desktop\CLE3_PT_Trends.xlsx'

# ── Palette ───────────────────────────────────────────────────
HDR_FILL = PatternFill('solid', fgColor='1F3864')
SEC_FILL = PatternFill('solid', fgColor='2E4D8A')
ALT_FILL = PatternFill('solid', fgColor='EEF2FF')
FAC_FILL = PatternFill('solid', fgColor='2E4D8A')
thin = Border(
    left  =Side(style='thin', color='C0C0C0'),
    right =Side(style='thin', color='C0C0C0'),
    top   =Side(style='thin', color='C0C0C0'),
    bottom=Side(style='thin', color='C0C0C0'))

# Shift definitions: (key, filter fn, display label, line color)
SHIFTS = [
    ('Night', lambda s: 'Night' in s,     'Night Shift (6p–6a)', '2E4D8A'),
    ('Day',   lambda s: 'Day Shift' in s, 'Day Shift (6a–6p)',   '16A34A'),
    ('All',   lambda s: s.strip()=='All Day', 'All Day',          'FF9900'),
]

COLORS_AM = [
    'E41A1C','377EB8','4DAF4A','984EA3','FF7F00','A65628','F781BF',
    '66C2A5','FC8D62','8DA0CB','E78AC3','A6D854','FFD92F','1B9E77',
    'D95F02','7570B3','E7298A','66A61E','E6AB02','A6761D','8DD3C7',
    'FB8072','80B1D3','FDB462','B3DE69','BC80BD','BEBADA','FFFFB3',
    '999999','666666','B3B3B3','CCEBC5','FCCDE5','D9D9D9','FFED6F','E5C494']

# ── Helpers ───────────────────────────────────────────────────
def _hdr(ws, row, col, val):
    c = ws.cell(row=row, column=col, value=val)
    c.font = Font(name='Segoe UI', size=10, bold=True, color='FFFFFF')
    c.fill = HDR_FILL
    c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    c.border = thin
    return c

def _dat(ws, row, col, val, bold=False, fill=None, align='center'):
    c = ws.cell(row=row, column=col, value=val)
    c.font = Font(name='Segoe UI', size=10, bold=bold)
    if fill: c.fill = fill
    c.alignment = Alignment(horizontal=align, vertical='center')
    c.border = thin
    return c

def _banner(ws, row, text, ncols, big=False):
    end_col = max(ncols, 4)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=end_col)
    c = ws.cell(row, 1, text)
    c.font = Font(name='Segoe UI', size=13 if big else 11, bold=True, color='FFFFFF')
    c.fill = HDR_FILL if big else SEC_FILL
    c.alignment = Alignment(horizontal='center' if big else 'left', vertical='center')
    ws.row_dimensions[row].height = 32 if big else 26

def _to_date(v):
    if isinstance(v, datetime.datetime): return v.date()
    if isinstance(v, str): return datetime.datetime.strptime(v[:10], '%Y-%m-%d').date()
    return v

def _norm_date_key(v):
    """Normalize a date value to a 'YYYY-MM-DD' string for dict keys."""
    if isinstance(v, datetime.datetime): return v.strftime('%Y-%m-%d')
    if isinstance(v, datetime.date):     return v.strftime('%Y-%m-%d')
    if isinstance(v, str):               return v[:10]
    return str(v)[:10]


# ── Tab 1: Overall PT% Trend ──────────────────────────────────
def build_overall_tab(wb, shift_data):
    """
    shift_data: dict keyed by shift key ('Night','Day','All'),
                value = {date_key: pt_value}
    """
    ws = wb.create_sheet('Overall PT% Trend')
    ws.sheet_view.showGridLines = False

    # Union of all dates, sorted
    all_date_keys = sorted(set(
        dk for sd in shift_data.values() for dk in sd.keys()
    ))

    if not all_date_keys:
        ws.cell(2, 1, 'No data in history yet.')
        return

    _banner(ws, 1, 'CLE3  ·  Overall Productive Time %  ·  All Shifts', 6, big=True)

    # Headers: Date | Night | Day | All Day | Target
    headers = ['Date', 'Night Shift', 'Day Shift', 'All Day', 'Target (84%)']
    col_widths = [12, 14, 14, 14, 14]
    for ci, (h, w) in enumerate(zip(headers, col_widths), 1):
        _hdr(ws, 2, ci, h)
        ws.column_dimensions[get_column_letter(ci)].width = w

    n = len(all_date_keys)
    for i, dk in enumerate(all_date_keys):
        r, f = i + 3, ALT_FILL if i % 2 == 0 else None
        c = _dat(ws, r, 1, _to_date(dk), fill=f)
        c.number_format = 'MM/DD'
        for ci, key in enumerate(['Night', 'Day', 'All'], 2):
            v = shift_data[key].get(dk)
            cell = _dat(ws, r, ci, v, fill=f)
            cell.number_format = '0.0'
        _dat(ws, r, 5, 84, fill=f)

    # Chart
    ch = LineChart()
    ch.title = 'CLE3 Overall Productive Time %'
    ch.style = 10
    ch.y_axis.title = 'PT%'; ch.x_axis.title = 'Date'
    ch.y_axis.scaling.min = 60; ch.y_axis.scaling.max = 100
    ch.height, ch.width = 16, 26

    shift_line_colors = ['2E4D8A', '16A34A', 'FF9900']
    for ci, (key, _, label, color) in enumerate(SHIFTS, 2):
        ref = Reference(ws, min_col=ci, min_row=2, max_row=2 + n)
        ch.add_data(ref, titles_from_data=True)
        si = ci - 2
        ch.series[si].graphicalProperties.line.solidFill = color
        ch.series[si].graphicalProperties.line.width = 20000
        ch.series[si].marker.symbol = 'circle'; ch.series[si].marker.size = 6
        ch.series[si].marker.graphicalProperties.solidFill = color

    # 84% target line
    tgt_ref = Reference(ws, min_col=5, min_row=2, max_row=2 + n)
    ch.add_data(tgt_ref, titles_from_data=True)
    si = 3
    ch.series[si].graphicalProperties.line.solidFill = 'C00000'
    ch.series[si].graphicalProperties.line.width = 15000
    ch.series[si].graphicalProperties.line.dashDot = 'dash'
    ch.series[si].marker.symbol = 'none'

    ch.set_categories(Reference(ws, min_col=1, min_row=3, max_row=2 + n))
    ws.add_chart(ch, 'G2')


# ── Tab 2: AM PT% Trends ──────────────────────────────────────
def build_am_tab(wb, all_daily, all_mgr):
    ws = wb.create_sheet('AM PT% Trends')
    ws.sheet_view.showGridLines = False
    ws.column_dimensions['A'].width = 26

    _banner(ws, 1, 'CLE3  ·  Area Manager PT% Trends  ·  All Shifts', 10, big=True)

    current_row = 3   # start after title

    for shift_key, shift_filt, shift_title, shift_color in SHIFTS:
        daily = [r for r in all_daily if shift_filt(str(r[1] or ''))]
        mgr_r = [r for r in all_mgr   if shift_filt(str(r[1] or ''))]

        if not daily:
            _banner(ws, current_row, f'▸  {shift_title}  —  no data yet', 4)
            current_row += 2
            continue

        dates   = [r[0] for r in daily]
        pt_vals = [r[2] for r in daily]

        mgr_data = defaultdict(dict)
        for row in mgr_r:
            if row[2] and row[3] is not None:
                mgr_data[row[2]][row[0]] = round(row[3], 1)
        managers = sorted(mgr_data.keys())

        ncols = 1 + len(dates)
        _banner(ws, current_row, f'▸  {shift_title}', ncols)
        current_row += 1

        # Header row: Area Manager | date1 | date2 | ...
        _hdr(ws, current_row, 1, 'Area Manager')
        for ci, d in enumerate(dates):
            col = ci + 2
            c = _hdr(ws, current_row, col, _to_date(d))
            c.number_format = 'MM/DD'
            ws.column_dimensions[get_column_letter(col)].width = 9
        hdr_row = current_row
        current_row += 1

        # Facility overall row
        fac_r = current_row
        c = ws.cell(fac_r, 1, 'FACILITY OVERALL')
        c.font = Font(name='Segoe UI', size=10, bold=True, color='FFFFFF')
        c.fill = FAC_FILL
        c.alignment = Alignment(horizontal='left', vertical='center'); c.border = thin
        for ci, v in enumerate(pt_vals):
            cell = ws.cell(fac_r, ci+2, v)
            cell.font = Font(name='Segoe UI', size=10, bold=True, color='FFFFFF')
            cell.fill = FAC_FILL; cell.number_format = '0.0'
            cell.alignment = Alignment(horizontal='center', vertical='center'); cell.border = thin
        current_row += 1

        # Manager rows
        for mi, mgr in enumerate(managers):
            f = ALT_FILL if mi % 2 == 0 else None
            _dat(ws, current_row, 1, mgr, fill=f, align='left')
            for ci, d in enumerate(dates):
                cell = _dat(ws, current_row, ci+2, mgr_data[mgr].get(d), fill=f)
                cell.number_format = '0.0'
            current_row += 1

        # AM chart
        ch = LineChart()
        ch.title = f'AM PT%  ·  {shift_title}'
        ch.style = 10
        ch.y_axis.title = 'PT%'; ch.x_axis.title = 'Date'
        ch.y_axis.scaling.min = 55; ch.y_axis.scaling.max = 100
        ch.height, ch.width = 18, 32

        fac_ref = Reference(ws, min_col=2, max_col=1+len(dates),
                            min_row=fac_r, max_row=fac_r)
        ch.add_data(fac_ref)
        ch.series[0].title = SeriesLabel(v='FACILITY')
        ch.series[0].graphicalProperties.line.solidFill = '000000'
        ch.series[0].graphicalProperties.line.width = 28000

        for mi, mgr in enumerate(managers):
            row = fac_r + 1 + mi
            ref = Reference(ws, min_col=2, max_col=1+len(dates), min_row=row, max_row=row)
            ch.add_data(ref)
            si = mi + 1
            ch.series[si].title = SeriesLabel(v=mgr.split(',')[0])
            color = COLORS_AM[mi % len(COLORS_AM)]
            ch.series[si].graphicalProperties.line.solidFill = color
            ch.series[si].graphicalProperties.line.width = 14000
            ch.series[si].marker.symbol = 'circle'; ch.series[si].marker.size = 4
            ch.series[si].marker.graphicalProperties.solidFill = color

        ch.set_categories(Reference(ws, min_col=2, max_col=1+len(dates),
                                     min_row=hdr_row, max_row=hdr_row))
        ws.add_chart(ch, f'A{current_row + 1}')
        current_row += 26   # space for chart + gap between shifts


# ── Tab 3: Flagged Count Trend ────────────────────────────────
def build_flagged_tab(wb, shift_data_flagged, shift_data_total):
    """
    shift_data_flagged / shift_data_total: dicts keyed by shift key,
    value = {date_key: count}
    """
    ws = wb.create_sheet('Flagged Count Trend')
    ws.sheet_view.showGridLines = False

    all_date_keys = sorted(set(
        dk for sd in shift_data_flagged.values() for dk in sd.keys()
    ))

    if not all_date_keys:
        ws.cell(2, 1, 'No data in history yet.')
        return

    _banner(ws, 1, 'CLE3  ·  Associates Flagged Below 84%  ·  All Shifts', 7, big=True)

    headers = ['Date',
               'Night Flagged', 'Day Flagged', 'All Day Flagged',
               'Night Total',   'Day Total',   'All Day Total']
    col_widths = [12, 15, 13, 16, 13, 11, 14]
    for ci, (h, w) in enumerate(zip(headers, col_widths), 1):
        _hdr(ws, 2, ci, h)
        ws.column_dimensions[get_column_letter(ci)].width = w

    n = len(all_date_keys)
    for i, dk in enumerate(all_date_keys):
        r, f = i + 3, ALT_FILL if i % 2 == 0 else None
        c = _dat(ws, r, 1, _to_date(dk), fill=f)
        c.number_format = 'MM/DD'
        for ci, key in enumerate(['Night', 'Day', 'All'], 2):
            _dat(ws, r, ci, shift_data_flagged[key].get(dk), fill=f)
        for ci, key in enumerate(['Night', 'Day', 'All'], 5):
            _dat(ws, r, ci, shift_data_total[key].get(dk), fill=f)

    # Chart — flagged count lines
    ch = LineChart()
    ch.title = 'Associates Flagged Below 84%'
    ch.style = 10
    ch.y_axis.title = 'Number of Associates'; ch.x_axis.title = 'Date'
    ch.height, ch.width = 16, 26

    line_colors = ['C00000', '16A34A', 'FF9900']
    for ci, (key, _, label, color) in enumerate(SHIFTS, 2):
        ref = Reference(ws, min_col=ci, min_row=2, max_row=2 + n)
        ch.add_data(ref, titles_from_data=True)
        si = ci - 2
        ch.series[si].graphicalProperties.line.solidFill = line_colors[si]
        ch.series[si].graphicalProperties.line.width = 20000
        ch.series[si].marker.symbol = 'circle'; ch.series[si].marker.size = 6
        ch.series[si].marker.graphicalProperties.solidFill = line_colors[si]

    ch.set_categories(Reference(ws, min_col=1, min_row=3, max_row=2 + n))
    ws.add_chart(ch, 'I2')


# ── Main ──────────────────────────────────────────────────────
def main():
    src       = openpyxl.load_workbook(SRC_PATH)
    all_daily = list(src['Daily Summary'].iter_rows(min_row=2, values_only=True))
    all_mgr   = list(src['Manager Trends'].iter_rows(min_row=2, values_only=True))

    # Build per-shift lookup dicts
    shift_daily_pt   = {k: {} for k, *_ in SHIFTS}
    shift_flagged    = {k: {} for k, *_ in SHIFTS}
    shift_total      = {k: {} for k, *_ in SHIFTS}

    for row in all_daily:
        shift_str = str(row[1] or '')
        for key, filt, *_ in SHIFTS:
            if filt(shift_str):
                dk = _norm_date_key(row[0])
                shift_daily_pt[key][dk] = row[2]   # overall PT%
                shift_flagged[key][dk]  = row[4]   # flagged count
                shift_total[key][dk]    = row[3]   # total associates
                break

    wb = openpyxl.Workbook()
    wb.remove(wb.active)   # remove default blank sheet

    build_overall_tab(wb, shift_daily_pt)
    build_am_tab(wb, all_daily, all_mgr)
    build_flagged_tab(wb, shift_flagged, shift_total)

    wb.save(OUT_PATH)
    print(f'Saved: {OUT_PATH}')


if __name__ == '__main__':
    main()
